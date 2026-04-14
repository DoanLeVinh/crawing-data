import csv
import random
import re
import unicodedata
from datetime import datetime, timedelta
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, Side

INPUT_FILES = [
	"Final_Master_Data_v3_cleaned.csv",
]
RAW_MASTER_FILE = "Final_Master_Data_v3.csv"
OUTPUT_FILE = "Invoices_11000_ReceiptStyle_OneSheet.xlsx"
OUTPUT_CSV_FILE = "Invoices_11000_ReceiptStyle_OneSheet.csv"
SHEET_NAME = "Receipts"
TOTAL_INVOICES = 11000
SEED = 42
VAT_RATES = [0.08, 0.10]
ENABLE_LOTTE_DOWNSAMPLING = True

STORE_INFO = {
	"storeName": "BACH HOA XANH",
	"website": "www.bachhoaxanh.com",
	"storeAddress": "49 - 49A Huynh Tinh Cua, Phuong Vo Thi Sau, Quan 3, Thanh pho Ho Chi Minh, Viet Nam",
}

CASHIERS = [
	"Nguyen Thi Thu Hien",
	"Tran Minh Chau",
	"Le Quoc Bao",
	"Pham Hoang Long",
	"Vu Ngoc Anh",
	"Do Thanh Truc",
	"Bui Tuan Kiet",
]

TITLE_FONT = Font(size=14, bold=True)
HEADER_FONT = Font(size=12, bold=True)
BOLD_FONT = Font(bold=True)
CENTER_ALIGN = Alignment(horizontal="center")
RIGHT_ALIGN = Alignment(horizontal="right")
WRAP_CENTER = Alignment(horizontal="center", wrap_text=True)
BOTTOM_BORDER = Border(bottom=Side(style="thin", color="000000"))


def to_int(value, default=0):
	try:
		if value is None or value == "":
			return default
		return int(float(value))
	except Exception:
		return default


def clean_product_name(name):
	if not name:
		return "San pham"
	return re.sub(r"\s+", " ", name).strip()


def clean_text_field(value, to_lower=False, default=""):
	if value is None:
		text = ""
	else:
		text = str(value)
	text = text.replace("\ufeff", " ").replace("\u00a0", " ").replace("\ufffd", " ")
	text = re.sub(r"[\x00-\x1f\x7f]+", " ", text)
	text = re.sub(r"\s+", " ", text).strip()
	if to_lower:
		text = text.lower()
	if not text:
		return default
	return text


def clean_label_field(value, default="unknown"):
	text = clean_text_field(value, to_lower=True, default="")
	if not text:
		return default
	text = re.sub(r"[^\w\s\-\./&+]", "", text, flags=re.UNICODE)
	text = re.sub(r"\s+", " ", text).strip(" -./")
	if not text:
		return default
	if not any(ch.isalpha() for ch in text):
		return default
	return text[:80]


def clean_price_value(value):
	parsed = to_int(clean_text_field(value, default="0"), default=0)
	if parsed < 0:
		return 0
	return parsed


def write_clean_master_csv(raw_path, cleaned_path):
	if not raw_path.exists():
		raise FileNotFoundError(f"Khong tim thay file master goc: {raw_path}")

	cleaned_rows = []
	with raw_path.open("r", encoding="utf-8-sig", newline="") as f:
		reader = csv.DictReader(f)
		for row in reader:
			pid = clean_text_field(row.get("product_id") or row.get("\ufeffproduct_id"), default="")
			if not pid:
				continue

			cleaned_rows.append(
				{
					"product_id": pid,
					"product_name": clean_text_field(row.get("product_name"), to_lower=True, default="san pham"),
					"brand": clean_label_field(row.get("brand"), default="unknown"),
					"category": clean_text_field(row.get("category"), default=""),
					"original_price": str(clean_price_value(row.get("original_price"))),
					"sale_price": str(clean_price_value(row.get("sale_price"))),
					"rating_count": str(max(0, to_int(row.get("rating_count"), default=0))),
					"frequently_bought": clean_text_field(row.get("frequently_bought"), to_lower=True, default=""),
					"shop_type": clean_text_field(row.get("shop_type"), default=""),
					"discount_percent": clean_text_field(row.get("discount_percent"), default=""),
					"platform": clean_label_field(row.get("platform"), default="unknown"),
					"thumbnail_url": clean_text_field(row.get("thumbnail_url"), default=""),
					"source_platform": clean_label_field(row.get("source_platform"), default="unknown"),
					"product_url": clean_text_field(row.get("product_url"), default=""),
					"crawl_timestamp": clean_text_field(row.get("crawl_timestamp"), default=""),
				}
			)

	fieldnames = [
		"product_id",
		"product_name",
		"brand",
		"category",
		"original_price",
		"sale_price",
		"rating_count",
		"frequently_bought",
		"shop_type",
		"discount_percent",
		"platform",
		"thumbnail_url",
		"source_platform",
		"product_url",
		"crawl_timestamp",
	]

	with cleaned_path.open("w", encoding="utf-8-sig", newline="") as f:
		writer = csv.DictWriter(f, fieldnames=fieldnames)
		writer.writeheader()
		writer.writerows(cleaned_rows)

	return len(cleaned_rows)


def normalize_platform(platform, source_platform, shop_type):
	def canonicalize_platform(raw):
		p = normalize_text(raw)
		if not p:
			return ""

		# Gop tat ca bien the WinMart ve 1 nhan duy nhat de tranh tach nen tang.
		if "winmart" in p:
			return "winmart"
		return p

	pl = normalize_text(platform)
	sp = normalize_text(source_platform)
	st = normalize_text(shop_type)

	if pl:
		return canonicalize_platform(pl)
	if sp:
		return canonicalize_platform(sp)
	if st:
		return canonicalize_platform(st)
	return "unknown"


def normalize_text(text):
	if not text:
		return ""
	txt = text.lower().strip()
	txt = unicodedata.normalize("NFD", txt)
	txt = "".join(ch for ch in txt if unicodedata.category(ch) != "Mn")
	return re.sub(r"\s+", " ", txt)


def tokenize(text):
	return re.findall(r"[a-z0-9]+", normalize_text(text))


def get_base_product_id(pid):
	return re.sub(r"_v\d+$", "", (pid or "").strip())


def infer_category(product_name, raw_category):
	text = normalize_text(product_name)
	cat = normalize_text(raw_category)
	tokens = set(tokenize(product_name))

	if any(k in text for k in ["ca rot", "ca chua", "bap cai"]) or any(t in tokens for t in ["rau", "hanh", "toi"]):
		return "Rau"
	if any(k in text for k in ["hai san"]) or any(t in tokens for t in ["thit", "ga", "bo", "heo"]):
		return "Thit"
	if any(k in text for k in ["nuoc", "tra", "sua", "sting", "pepsi", "fanta", "coca", "soda", "bia"]):
		return "DoUong"
	if "fmcg" in cat:
		return "DoUong"
	return "Khac"


def is_machine_product(product_name, raw_category):
	name = normalize_text(product_name)
	cat = normalize_text(raw_category)

	# Loai cac mat hang co tu khoa "may"/"machine" de tranh dua thiet bi vao don hang.
	if re.search(r"\bmay\b", name) is not None:
		return True
	if "machine" in name:
		return True
	if any(k in cat for k in ["dien gia dung", "do dien", "dien may"]):
		return True
	return False


def robust_price(prices):
	vals = sorted([p for p in prices if p > 0])
	if not vals:
		return 0
	n = len(vals)
	if n % 2:
		return vals[n // 2]
	return int(round((vals[n // 2 - 1] + vals[n // 2]) / 2))


def normalize_prices_scale_x10(prices):
	vals = [int(p) for p in prices if p and p > 0]
	if len(vals) < 2:
		return vals

	mn = min(vals)
	mx = max(vals)
	ratio = mx / mn if mn else 0

	# Neu cung 1 san pham co cum gia lech khoang x10, dua cum cao ve cung scale.
	if not (8 <= ratio <= 12):
		return vals

	normalized = []
	for p in vals:
		if p >= mn * 7:
			normalized.append(int(round(p / 10)))
		else:
			normalized.append(p)
	return normalized


def normalize_price_by_category(price, category, product_name):
	if price <= 0:
		return 0

	# Rule theo yeu cau: gia vuot 1,000,000 thi dat ve 100,000.
	if price > 1000000:
		return 100000

	# Rule theo yeu cau: gia vuot 250000 thi quy ve tien chuc (chia 10) den khi <= 250000.
	adjusted = int(price)
		
	while adjusted > 250000:
		adjusted = int(round(adjusted / 10))

	name = normalize_text(product_name)
	max_price = {
		"DoUong": 500000,
		"Rau": 400000,
		"Thit": 1500000,
		"Khac": 2000000,
	}.get(category, 2000000)

	# Mat ong, gia vi va cac san pham dong goi nho thuong khong vuot qua nguong nay.
	if "mat ong" in name:
		max_price = min(max_price, 500000)

	# Nhom binh xit gia dung thuong o muc thap, can chan loi scale x10.
	if any(k in name for k in ["binh xit", "xit con trung", "xit muoi", "vape", "ars jet"]):
		max_price = min(max_price, 180000)

	while adjusted > max_price:
		adjusted = int(round(adjusted / 10))

	return max(1000, adjusted)


def load_products_from_sources(input_paths):
	merged = {}

	for path in input_paths:
		with path.open("r", encoding="utf-8", newline="") as f:
			reader = csv.DictReader(f)
			for row in reader:
				name = clean_product_name(row.get("product_name", "")).lower()
				brand = clean_product_name(row.get("brand", "")).lower()
				raw_cat = row.get("category", "")
				if is_machine_product(name, raw_cat):
					continue
				sale_price = to_int(row.get("sale_price"), default=0)
				rating = max(1, to_int(row.get("rating_count"), default=1))
				if sale_price <= 0:
					continue

				pid = (row.get("product_id", "") or row.get("\ufeffproduct_id", "") or "").strip()
				platform = normalize_platform(row.get("platform", ""), row.get("source_platform", ""), row.get("shop_type", ""))
				base_key = get_base_product_id(pid) if pid else normalize_text(name)
				key = f"{platform}::{base_key}"

				entry = merged.setdefault(
					key,
					{
						"product_id": key,
						"platform": platform,
						"name_counts": {},
						"brand_counts": {},
						"prices": [],
						"rating_count": 0,
						"categories": [],
						"frequently_bought": set(),
					},
				)

				entry["name_counts"][name] = entry["name_counts"].get(name, 0) + 1
				if brand:
					entry["brand_counts"][brand] = entry["brand_counts"].get(brand, 0) + 1
				entry["prices"].append(sale_price)
				entry["rating_count"] = max(entry["rating_count"], rating)
				entry["categories"].append(raw_cat)

				for fb in (row.get("frequently_bought", "") or "").split(","):
					fb_name = clean_product_name(fb).lower()
					if fb_name:
						entry["frequently_bought"].add(fb_name)

	products = []
	for value in merged.values():
		best_name = sorted(value["name_counts"].items(), key=lambda x: (x[1], len(x[0])), reverse=True)[0][0]
		best_brand = ""
		if value["brand_counts"]:
			best_brand = sorted(value["brand_counts"].items(), key=lambda x: (x[1], len(x[0])), reverse=True)[0][0]
		raw_cat = value["categories"][0] if value["categories"] else ""
		category = infer_category(best_name, raw_cat)
		scaled_prices = normalize_prices_scale_x10(value["prices"])
		unit_price = robust_price(scaled_prices)
		unit_price = normalize_price_by_category(unit_price, category, best_name)
		if unit_price <= 0:
			continue

		products.append(
			{
				"product_id": value["product_id"],
				"product_name": best_name,
				"brand": best_brand,
				"unitPrice": unit_price,
				"rating_count": value["rating_count"],
				"category": category,
				"platform": value["platform"],
				"frequently_bought": sorted(value["frequently_bought"]),
			}
		)

	if not products:
		raise ValueError("Khong doc duoc san pham hop le tu file dau vao")
	return products


def build_name_index(products):
	return {normalize_text(p["product_name"]): p for p in products}


def build_copurchase_graph(products):
	name_index = build_name_index(products)
	by_id = {p["product_id"]: p for p in products}
	graph = {p["product_id"]: set() for p in products}

	for p in products:
		pid = p["product_id"]
		for fb_name in p["frequently_bought"]:
			fb = name_index.get(normalize_text(fb_name))
			if not fb:
				continue
			nid = fb["product_id"]
			if nid != pid and nid in by_id:
				graph[pid].add(nid)
				graph[nid].add(pid)
	return graph


def pick_random_product(products, used_ids):
	candidates = [p for p in products if p["product_id"] not in used_ids]
	if not candidates:
		return None
	return weighted_pick(candidates)


def is_weight_product(product):
	name = normalize_text(product["product_name"])

	# Cac tu khoa dong goi cho thay san pham ban theo don vi thay vi can ky.
	pack_keywords = ["hop", "goi", "chai", "lon", "thung", "bich", "tuyp", "combo", "set", "vi", "pack", "can", "tui"]
	if any(k in name for k in pack_keywords):
		return False

	if any(k in name for k in ["(kg)", " kg", "kilogram"]):
		return True

	if product["category"] == "Rau":
		return True

	fresh_meat_keywords = ["thit", "bo", "heo", "ga", "ca ", "hai san", "tom", "muc"]
	if product["category"] == "Thit" and any(k in name for k in fresh_meat_keywords):
		return True

	return False


def random_quantity(product):
	# Theo yeu cau: toan bo so luong deu la so nguyen.
	if is_weight_product(product):
		return random.choices([1, 2, 3], weights=[70, 22, 8], k=1)[0]
	return random.choices([1, 2, 3, 4, 5], weights=[55, 25, 12, 6, 2], k=1)[0]


def weighted_pick(candidates):
	weights = [max(1, c["rating_count"]) for c in candidates]
	return random.choices(candidates, weights=weights, k=1)[0]


def build_line_items(products, product_lookup, copurchase_graph):
	used_ids = set()
	selected = []
	target_count = random.randint(5, 12)
	pair_target = max(1, int(round(target_count * 0.6)))

	anchor = weighted_pick(products)
	for _ in range(8):
		if copurchase_graph.get(anchor["product_id"], set()):
			break
		anchor = weighted_pick(products)
	selected.append(anchor)
	used_ids.add(anchor["product_id"])

	# Buoc 1: uu tien bo sung san pham theo frequently_bought de tao cap dong mua.
	while len(selected) < pair_target:
		neighbor_ids = []
		for s in selected:
			neighbor_ids.extend(list(copurchase_graph.get(s["product_id"], set())))
		neighbor_candidates = [product_lookup[nid] for nid in neighbor_ids if nid in product_lookup and nid not in used_ids]
		if not neighbor_candidates:
			break
		candidate = weighted_pick(neighbor_candidates)
		selected.append(candidate)
		used_ids.add(candidate["product_id"])

	# Buoc 2: chen them san pham ngau nhien de mo phong mua sam tu do.
	while len(selected) < target_count:
		candidate = pick_random_product(products, used_ids)
		if candidate is None:
			break
		selected.append(candidate)
		used_ids.add(candidate["product_id"])

	subtotal = 0
	line_items = []
	for p in selected:
		quantity = random_quantity(p)
		unit_price = p["unitPrice"]
		line_amount = int(round(quantity * unit_price))
		subtotal += line_amount
		line_items.append(
			{
				"productName": p["product_name"],
				"brand": p.get("brand", ""),
				"quantity": quantity,
				"unitPrice": unit_price,
				"lineAmount": line_amount,
			}
		)

	return line_items, subtotal


def round_amount(amount):
	return int(round(amount / 500.0) * 500)


def cash_given(amount_due):
	base = int((amount_due + 999) // 1000 * 1000)
	return base + random.choice([0, 1000, 2000, 5000, 10000, 20000, 50000])


def format_qty(quantity):
	if abs(quantity - int(quantity)) < 1e-9:
		return str(int(quantity))
	return f"{quantity:.3f}"


def format_money(value):
	return f"{int(round(value)):,}"


def build_receipt(idx, products, product_lookup, copurchase_graph):
	receipt_datetime = datetime(2025, 1, 1) + timedelta(
		days=random.randint(0, 364),
		hours=random.randint(6, 21),
		minutes=random.randint(0, 59),
	)
	receipt_id = f"{receipt_datetime.strftime('%y%m%d')}{idx:011d}"

	line_items, subtotal_amount = build_line_items(products, product_lookup, copurchase_graph)
	vat_rate = random.choice(VAT_RATES)
	vat_amount = int(round(subtotal_amount * vat_rate))
	grand_total = subtotal_amount + vat_amount
	final_amount = round_amount(grand_total)
	given = cash_given(final_amount)

	return {
		"storeName": STORE_INFO["storeName"],
		"website": STORE_INFO["website"],
		"storeAddress": STORE_INFO["storeAddress"],
		"receiptId": receipt_id,
		"receiptDateTime": receipt_datetime.strftime("%d/%m/%Y %H:%M"),
		"cashierName": random.choice(CASHIERS),
		"lineItems": line_items,
		"subtotalAmount": subtotal_amount,
		"vatRate": vat_rate,
		"vatAmount": vat_amount,
		"totalAmount": grand_total,
		"finalAmount": final_amount,
		"cashGiven": given,
		"changeReturned": given - final_amount,
	}


def build_platform_quotas(products, total_invoices):
	platforms = sorted({p["platform"] for p in products if p.get("platform")})
	if not platforms:
		raise ValueError("Khong tim thay thong tin san/platform hop le trong du lieu")

	base = total_invoices // len(platforms)
	remainder = total_invoices % len(platforms)
	quotas = {pl: base for pl in platforms}
	for i in range(remainder):
		quotas[platforms[i]] += 1
	return quotas


def is_lotte_platform(platform_name):
	pl = normalize_text(platform_name)
	return "lotte" in pl


def downsample_lotte_products(products):
	by_platform = {}
	for p in products:
		by_platform.setdefault(p["platform"], []).append(p)

	non_lotte_counts = [len(pool) for pl, pool in by_platform.items() if not is_lotte_platform(pl)]
	if not non_lotte_counts:
		return products, {}

	# Dua quy mo lotte ve muc tuong duong cac san con lai de tranh bias theo san.
	target_size = max(1, int(round(sum(non_lotte_counts) / len(non_lotte_counts))))
	balanced = []
	stats = {}

	for pl, pool in by_platform.items():
		before = len(pool)
		if is_lotte_platform(pl) and before > target_size:
			pool = random.sample(pool, target_size)
		after = len(pool)
		stats[pl] = {"before": before, "after": after}
		balanced.extend(pool)

	return balanced, stats


def build_platform_context(products):
	by_platform = {}
	for p in products:
		by_platform.setdefault(p["platform"], []).append(p)

	context = {}
	for pl, pool in by_platform.items():
		context[pl] = {
			"products": pool,
			"lookup": {x["product_id"]: x for x in pool},
			"graph": build_copurchase_graph(pool),
		}
	return context


def build_platform_schedule(quotas):
	schedule = []
	for pl, cnt in quotas.items():
		schedule.extend([pl] * cnt)
	random.shuffle(schedule)
	return schedule


def receipt_to_csv_row(receipt):
	item_names = []
	item_brands = []
	item_qty = []
	item_prices = []
	item_amounts = []
	for item in receipt["lineItems"]:
		item_names.append(item["productName"])
		item_brands.append(item.get("brand", ""))
		item_qty.append(str(item["quantity"]))
		item_prices.append(str(item["unitPrice"]))
		item_amounts.append(str(item["lineAmount"]))

	transaction_items = ",".join(item_names)
	transaction_brands = ",".join([x for x in item_brands if x])

	return {
		"platform": clean_label_field(receipt.get("platform", "unknown"), default="unknown"),
		"storeName": receipt["storeName"],
		"website": receipt["website"],
		"storeAddress": receipt["storeAddress"],
		"receiptId": receipt["receiptId"],
		"receiptDateTime": receipt["receiptDateTime"],
		"cashierName": receipt["cashierName"],
		"itemCount": len(receipt["lineItems"]),
		"productNames": " | ".join(item_names),
		"brands": " | ".join(item_brands),
		"quantities": " | ".join(item_qty),
		"unitPrices": " | ".join(item_prices),
		"lineAmounts": " | ".join(item_amounts),
		"transactionItems": transaction_items,
		"transactionBrands": transaction_brands,
		"subtotalAmount": receipt["subtotalAmount"],
		"vatRate": receipt["vatRate"],
		"vatAmount": receipt["vatAmount"],
		"totalAmount": receipt["totalAmount"],
		"finalAmount": receipt["finalAmount"],
		"cashGiven": receipt["cashGiven"],
		"changeReturned": receipt["changeReturned"],
	}


def receipt_to_ready_rows(receipt):
	rows = []
	platform = clean_label_field(receipt.get("platform", "unknown"), default="unknown")
	for idx, item in enumerate(receipt["lineItems"], start=1):
		brand = clean_label_field(item.get("brand", "unknown"), default="unknown")
		rows.append(
			{
				"receiptId": receipt["receiptId"],
				"receiptDateTime": receipt["receiptDateTime"],
				"platform": platform,
				"itemIndex": idx,
				"productName": clean_text_field(item["productName"], to_lower=True, default="san pham"),
				"brand": brand,
				"quantity": item["quantity"],
				"unitPrice": item["unitPrice"],
				"lineAmount": item["lineAmount"],
				"subtotalAmount": receipt["subtotalAmount"],
				"vatRate": receipt["vatRate"],
				"vatAmount": receipt["vatAmount"],
				"totalAmount": receipt["totalAmount"],
				"finalAmount": receipt["finalAmount"],
			}
		)
	return rows


def build_label_mapping(values):
	labels = sorted({clean_label_field(v, default="unknown") for v in values})
	return {label: idx for idx, label in enumerate(labels)}


def apply_label_encoding(ready_rows):
	platform_map = build_label_mapping([r["platform"] for r in ready_rows])
	brand_map = build_label_mapping([r["brand"] for r in ready_rows])
	default_platform_id = platform_map.get("unknown", 0)
	default_brand_id = brand_map.get("unknown", 0)

	for r in ready_rows:
		r["platform_id"] = platform_map.get(r["platform"], default_platform_id)
		r["brand_id"] = brand_map.get(r["brand"], default_brand_id)

	return platform_map, brand_map


def write_mapping_csv(mapping, out_path, key_col, id_col):
	with out_path.open("w", encoding="utf-8-sig", newline="") as f:
		writer = csv.DictWriter(f, fieldnames=[key_col, id_col])
		writer.writeheader()
		for label, idx in sorted(mapping.items(), key=lambda x: x[1]):
			writer.writerow({key_col: label, id_col: idx})


def write_ready_stage3_csv(ready_rows, output_path):
	fieldnames = [
		"receiptId",
		"receiptDateTime",
		"platform",
		"platform_id",
		"itemIndex",
		"productName",
		"brand",
		"brand_id",
		"quantity",
		"unitPrice",
		"lineAmount",
		"subtotalAmount",
		"vatRate",
		"vatAmount",
		"totalAmount",
		"finalAmount",
	]
	with output_path.open("w", encoding="utf-8-sig", newline="") as f:
		writer = csv.DictWriter(f, fieldnames=fieldnames)
		writer.writeheader()
		writer.writerows(ready_rows)


def build_encoded_invoice_rows(ready_rows):
	encoded_rows = []
	for r in ready_rows:
		encoded_rows.append(
			{
				"receiptId": r["receiptId"],
				"receiptDateTime": r["receiptDateTime"],
				"platform": r["platform_id"],
				"platform_id": r["platform_id"],
				"itemIndex": r["itemIndex"],
				"productName": r["productName"],
				"brand": r["brand_id"],
				"brand_id": r["brand_id"],
				"quantity": r["quantity"],
				"unitPrice": r["unitPrice"],
				"lineAmount": r["lineAmount"],
				"subtotalAmount": r["subtotalAmount"],
				"vatRate": r["vatRate"],
				"vatAmount": r["vatAmount"],
				"totalAmount": r["totalAmount"],
				"finalAmount": r["finalAmount"],
				"platform_text": r["platform"],
				"brand_text": r["brand"],
			}
		)
	return encoded_rows


def write_encoded_invoice_csv(encoded_rows, output_path):
	if not encoded_rows:
		return
	fieldnames = list(encoded_rows[0].keys())
	with output_path.open("w", encoding="utf-8-sig", newline="") as f:
		writer = csv.DictWriter(f, fieldnames=fieldnames)
		writer.writeheader()
		writer.writerows(encoded_rows)


def write_encoded_invoice_excel(encoded_rows, output_path):
	wb = Workbook()
	ws = wb.active
	ws.title = "EncodedInvoices"

	if not encoded_rows:
		wb.save(output_path)
		return

	fieldnames = list(encoded_rows[0].keys())
	for col, name in enumerate(fieldnames, start=1):
		ws.cell(1, col, name)
		ws.cell(1, col).font = BOLD_FONT

	for row_idx, row in enumerate(encoded_rows, start=2):
		for col_idx, name in enumerate(fieldnames, start=1):
			ws.cell(row_idx, col_idx, row[name])

	for col_idx, name in enumerate(fieldnames, start=1):
		width = min(42, max(12, len(name) + 2))
		ws.column_dimensions[ws.cell(1, col_idx).column_letter].width = width

	wb.save(output_path)


def write_receipt_block(ws, start_row, receipt):
	ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=4)
	ws.cell(start_row, 1, receipt["storeName"]).font = TITLE_FONT
	ws.cell(start_row, 1).alignment = CENTER_ALIGN

	ws.merge_cells(start_row=start_row + 1, start_column=1, end_row=start_row + 1, end_column=4)
	ws.cell(start_row + 1, 1, receipt["website"]).alignment = CENTER_ALIGN

	ws.merge_cells(start_row=start_row + 2, start_column=1, end_row=start_row + 2, end_column=4)
	ws.cell(start_row + 2, 1, receipt["storeAddress"]).alignment = WRAP_CENTER

	ws.merge_cells(start_row=start_row + 4, start_column=1, end_row=start_row + 4, end_column=4)
	ws.cell(start_row + 4, 1, "PHIEU THANH TOAN").font = HEADER_FONT
	ws.cell(start_row + 4, 1).alignment = CENTER_ALIGN

	ws.cell(start_row + 6, 1, "So CT:")
	ws.merge_cells(start_row=start_row + 6, start_column=2, end_row=start_row + 6, end_column=4)
	ws.cell(start_row + 6, 2, receipt["receiptId"])

	ws.cell(start_row + 7, 1, "Ngay CT:")
	ws.merge_cells(start_row=start_row + 7, start_column=2, end_row=start_row + 7, end_column=4)
	ws.cell(start_row + 7, 2, receipt["receiptDateTime"])

	ws.cell(start_row + 8, 1, "Nhan vien:")
	ws.merge_cells(start_row=start_row + 8, start_column=2, end_row=start_row + 8, end_column=4)
	ws.cell(start_row + 8, 2, receipt["cashierName"])

	header_row = start_row + 10
	ws.cell(header_row, 1, "Hang hoa")
	ws.cell(header_row, 2, "SL")
	ws.cell(header_row, 3, "Gia ban")
	ws.cell(header_row, 4, "T.Tien")

	for c in range(1, 5):
		ws.cell(header_row, c).font = BOLD_FONT
		ws.cell(header_row, c).alignment = CENTER_ALIGN
		ws.cell(header_row, c).border = BOTTOM_BORDER

	row = header_row + 1
	for item in receipt["lineItems"]:
		ws.cell(row, 1, item["productName"])
		ws.cell(row, 2, format_qty(item["quantity"]))
		ws.cell(row, 3, format_money(item["unitPrice"]))
		ws.cell(row, 4, format_money(item["lineAmount"]))
		ws.cell(row, 2).alignment = RIGHT_ALIGN
		ws.cell(row, 3).alignment = RIGHT_ALIGN
		ws.cell(row, 4).alignment = RIGHT_ALIGN
		row += 1

	row += 1
	ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
	ws.cell(row, 1, "Tam tinh:").alignment = RIGHT_ALIGN
	ws.cell(row, 4, format_money(receipt["subtotalAmount"]))
	ws.cell(row, 4).alignment = RIGHT_ALIGN
	ws.cell(row, 4).font = BOLD_FONT

	row += 1
	ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
	ws.cell(row, 1, f"VAT ({int(receipt['vatRate'] * 100)}%):").alignment = RIGHT_ALIGN
	ws.cell(row, 4, format_money(receipt["vatAmount"]))
	ws.cell(row, 4).alignment = RIGHT_ALIGN

	row += 1
	ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
	ws.cell(row, 1, "Tong tien:").alignment = RIGHT_ALIGN
	ws.cell(row, 4, format_money(receipt["totalAmount"]))
	ws.cell(row, 4).alignment = RIGHT_ALIGN
	ws.cell(row, 4).font = BOLD_FONT

	row += 1
	ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
	ws.cell(row, 1, "Thanh toan (Da lam tron):").alignment = RIGHT_ALIGN
	ws.cell(row, 4, format_money(receipt["finalAmount"]))
	ws.cell(row, 4).alignment = RIGHT_ALIGN
	ws.cell(row, 4).font = BOLD_FONT

	row += 1
	ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
	ws.cell(row, 1, "Tien mat:").alignment = RIGHT_ALIGN
	ws.cell(row, 4, format_money(receipt["cashGiven"]))
	ws.cell(row, 4).alignment = RIGHT_ALIGN

	row += 1
	ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
	ws.cell(row, 1, "Tien thoi lai:").alignment = RIGHT_ALIGN
	ws.cell(row, 4, format_money(receipt["changeReturned"]))
	ws.cell(row, 4).alignment = RIGHT_ALIGN

	return row + 3


def write_workbook(output_path, platform_context, platform_schedule):
	wb = Workbook()
	ws = wb.active
	ws.title = SHEET_NAME

	# Merge cells tren so luong lon gay cham dang O(n^2); tat merge de xuat 10k nhanh hon.
	ws.merge_cells = lambda *args, **kwargs: None

	ws.column_dimensions["A"].width = 46
	ws.column_dimensions["B"].width = 10
	ws.column_dimensions["C"].width = 14
	ws.column_dimensions["D"].width = 14

	next_row = 1
	csv_rows = []
	ready_rows = []
	platform_invoice_count = {}
	for i, platform in enumerate(platform_schedule, start=1):
		ctx = platform_context[platform]
		receipt = build_receipt(i, ctx["products"], ctx["lookup"], ctx["graph"])
		receipt["platform"] = platform
		platform_invoice_count[platform] = platform_invoice_count.get(platform, 0) + 1
		next_row = write_receipt_block(ws, next_row, receipt)
		csv_rows.append(receipt_to_csv_row(receipt))
		ready_rows.extend(receipt_to_ready_rows(receipt))
		if i % 300 == 0:
			print(f"Da sinh {i}/{TOTAL_INVOICES} hoa don...")

	try:
		wb.save(output_path)
		excel_path = output_path
	except PermissionError:
		ts = datetime.now().strftime("%Y%m%d_%H%M%S")
		fallback = output_path.with_name(f"{output_path.stem}_{ts}{output_path.suffix}")
		wb.save(fallback)
		print(f"Canh bao: File goc dang mo/bi khoa, da luu sang file moi: {fallback}")
		excel_path = fallback

	csv_path = output_path.with_suffix(".csv")
	fieldnames = [
		"platform",
		"storeName",
		"website",
		"storeAddress",
		"receiptId",
		"receiptDateTime",
		"cashierName",
		"itemCount",
		"productNames",
		"brands",
		"quantities",
		"unitPrices",
		"lineAmounts",
		"transactionItems",
		"transactionBrands",
		"subtotalAmount",
		"vatRate",
		"vatAmount",
		"totalAmount",
		"finalAmount",
		"cashGiven",
		"changeReturned",
	]
	with csv_path.open("w", encoding="utf-8-sig", newline="") as f:
		writer = csv.DictWriter(f, fieldnames=fieldnames)
		writer.writeheader()
		writer.writerows(csv_rows)

	platform_map, brand_map = apply_label_encoding(ready_rows)
	ready_path = output_path.with_name(f"{output_path.stem}_ready_stage3.csv")
	platform_map_path = output_path.with_name(f"{output_path.stem}_mapping_platform.csv")
	brand_map_path = output_path.with_name(f"{output_path.stem}_mapping_brand.csv")
	encoded_csv_path = output_path.with_name(f"{output_path.stem}_encoded.csv")
	encoded_excel_path = output_path.with_name(f"{output_path.stem}_encoded.xlsx")

	write_ready_stage3_csv(ready_rows, ready_path)
	write_mapping_csv(platform_map, platform_map_path, "platform", "platform_id")
	write_mapping_csv(brand_map, brand_map_path, "brand", "brand_id")
	encoded_rows = build_encoded_invoice_rows(ready_rows)
	write_encoded_invoice_csv(encoded_rows, encoded_csv_path)
	write_encoded_invoice_excel(encoded_rows, encoded_excel_path)

	return (
		excel_path,
		csv_path,
		ready_path,
		platform_map_path,
		brand_map_path,
		encoded_csv_path,
		encoded_excel_path,
		platform_invoice_count,
	)


def main():
	random.seed(SEED)

	base_dir = Path(__file__).resolve().parent
	raw_master_path = base_dir / RAW_MASTER_FILE
	cleaned_master_path = base_dir / INPUT_FILES[0]
	cleaned_count = write_clean_master_csv(raw_master_path, cleaned_master_path)

	input_paths = [base_dir / p for p in INPUT_FILES if (base_dir / p).exists()]
	if not input_paths:
		raise FileNotFoundError("Khong tim thay file dau vao")

	products = load_products_from_sources(input_paths)
	if ENABLE_LOTTE_DOWNSAMPLING:
		products, downsample_stats = downsample_lotte_products(products)
	else:
		downsample_stats = {}

	platform_context = build_platform_context(products)
	quotas = build_platform_quotas(products, TOTAL_INVOICES)
	platform_schedule = build_platform_schedule(quotas)

	output_path = base_dir / OUTPUT_FILE
	(
		saved_path,
		csv_path,
		ready_path,
		platform_map_path,
		brand_map_path,
		encoded_csv_path,
		encoded_excel_path,
		platform_stats,
	) = write_workbook(
		output_path, platform_context, platform_schedule
	)

	print(f"Da tao xong {TOTAL_INVOICES} hoa don")
	print(f"Da lam sach file goc: {cleaned_master_path} ({cleaned_count} dong)")
	print(f"So san pham hop le tu master data: {len(products)}")
	if downsample_stats:
		print("Down-sampling theo san (truoc -> sau):")
		for pl in sorted(downsample_stats):
			st = downsample_stats[pl]
			print(f"- {pl}: {st['before']} -> {st['after']}")
	print(f"So san duoc lay mau: {len(platform_stats)}")
	for pl in sorted(platform_stats):
		print(f"- {pl}: {platform_stats[pl]} hoa don")
	print(f"File Excel: {saved_path}")
	print(f"File CSV: {csv_path}")
	print(f"File Ready Stage 3: {ready_path}")
	print(f"Mapping platform: {platform_map_path}")
	print(f"Mapping brand: {brand_map_path}")
	print(f"Hoa don ma hoa CSV: {encoded_csv_path}")
	print(f"Hoa don ma hoa Excel: {encoded_excel_path}")


if __name__ == "__main__":
	main()
